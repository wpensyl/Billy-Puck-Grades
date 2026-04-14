#!/usr/bin/env python3
"""
Ice Grade — One-command update script.

Usage:
  python update.py

Expects all source files in the same directory as this script:
  VALUE_DRIVE_25-26.xlsm
  2025-26_Database.xlsm
  Comp_Sheet.xlsx
  nhl-analytics-v3.html  (template — no data embedded)

Outputs to ./output/:
  players_fwd.json
  players_dman.json
  comp_fwd.json
  comp_dman.json
  nhl-analytics-v3-embedded.html  (ready to open or deploy)

For Netlify deployment, drag the output/ folder onto your Netlify site.
"""

import sys, json, math, re, os, subprocess
from pathlib import Path

# ── Check dependencies ────────────────────────────────────────────────────────
try:
    import openpyxl
    from scipy.stats import norm as scipy_norm
except ImportError:
    print("Installing required packages...")
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "scipy", "--break-system-packages", "-q"])
    import openpyxl
    from scipy.stats import norm as scipy_norm

BASE = Path(__file__).parent
OUT  = BASE / "output"
OUT.mkdir(exist_ok=True)

# ── Source files ──────────────────────────────────────────────────────────────
VALUE_DRIVE = BASE / "VALUE_DRIVE_25-26.xlsm"
DATABASE    = BASE / "2025-26_Database.xlsm"
COMP_SHEET  = BASE / "Comp_Sheet.xlsx"
HTML_TPL    = BASE / "nhl-analytics-v3.html"

for f, label in [(VALUE_DRIVE,"VALUE_DRIVE"), (DATABASE,"Database"), (COMP_SHEET,"Comp_Sheet"), (HTML_TPL,"HTML template")]:
    if not f.exists():
        print(f"ERROR: {label} not found at {f}")
        sys.exit(1)

print("Ice Grade — Update script")
print(f"Source dir: {BASE}")
print(f"Output dir: {OUT}")
print()

# ════════════════════════════════════════════════════════════════════════════════
# STEP 1: Extract players + progression + contract calculator
# ════════════════════════════════════════════════════════════════════════════════
print("Step 1/3 — Extracting player data from VALUE DRIVE + Database...")

def pct(v):
    if v is None or not isinstance(v, (int, float)) or math.isnan(v): return None
    return round(v * 100)

def dollars(v):
    if v is None or not isinstance(v, (int, float)): return None
    return round(v)

def fmt_toi(total, gp):
    if not total or not gp: return None
    pg = total / gp; m = int(pg); s = round((pg - m) * 60)
    if s == 60: m += 1; s = 0
    return f"{m}:{s:02d}"

def fmt_height(v):
    if v is None: return None
    f = int(v); i = round((v - f) * 12)
    return f"{f}'{i}\""

def safe_str(v):
    return str(v).strip() if v is not None else None

def build_header(row):
    r = {}
    for i, v in enumerate(row):
        if v is not None: r[str(v).strip()] = i
    return r

def is_valid(v):
    if v is None or not isinstance(v, (int, float)) or math.isnan(v): return False
    if str(v) == '#N/A': return False
    return True

def z_to_pct(z):
    if not is_valid(z): return None
    return round(scipy_norm.cdf(z) * 100)

SEASONS = ['2020-21','2021-22','2022-23','2023-24','2024-25','2025-26']

def extract_prog_fwd(ws_hist, ws_curr, player_col_hist=0, player_col_curr=1):
    """Extract FWD 5-track progression.
    ws_hist = FWD 25 (year-by-year Z scores, header row 7)
    ws_curr = 25-26 FWD (current year norms, header row 2)
    
    Tracks:
      overall:       hist Z start 308, curr col 192
      off_prod:      hist Z start 318, curr col 195
      def_prod:      hist Z start 328, curr col 198
      off_analytics: hist Z start 338, curr col 201
      def_analytics: hist Z start 348, curr col 204
    """
    # Build current-year norms from 25-26 FWD
    curr_norms = {}  # name -> {overall, off_prod, def_prod, off_analytics, def_analytics}
    for row in ws_curr.iter_rows(min_row=3, values_only=True):
        name = row[player_col_curr]
        if not name or not isinstance(name, str): continue
        curr_norms[name.strip()] = {
            'overall':       row[192] if len(row)>192 else None,
            'off_prod':      row[195] if len(row)>195 else None,
            'def_prod':      row[198] if len(row)>198 else None,
            'off_analytics': row[201] if len(row)>201 else None,
            'def_analytics': row[204] if len(row)>204 else None,
        }

    HIST_STARTS = {'overall':308,'off_prod':318,'def_prod':328,'off_analytics':338,'def_analytics':348}
    players = {}
    for row in ws_hist.iter_rows(min_row=8, values_only=True):
        name = row[player_col_hist]
        if not name or not isinstance(name, str): continue
        cn = curr_norms.get(name.strip(), {})
        prog = {}
        for key, start in HIST_STARTS.items():
            out = []
            for i, s in enumerate(SEASONS):
                z = row[start+i] if start+i < len(row) else None
                out.append({'season': s, 'pct': z_to_pct(z) if is_valid(z) and str(z)!='#N/A' else None})
            # Override current year with the correct norm from 25-26 FWD
            curr_val = cn.get(key)
            if curr_val is not None and is_valid(curr_val):
                out[-1]['pct'] = pct(curr_val)
            prog[key] = out
        players[name.strip()] = prog
    return players

def extract_prog_dman(ws_hist, ws_curr, player_col_hist=0, player_col_curr=1):
    """Extract DMAN 5-track progression.
    ws_hist = DMAN 24 (year-by-year Z scores, header row 3)
    ws_curr = 25-26 DMAN (current year norms, header row 2)
    
    Tracks (DMAN 24 Z start cols, offset -20 from FWD):
      overall:       Z 288, curr col 187 (Overall Norm)
      off_prod:      Z 298, curr col 190
      def_prod:      Z 308, curr col 193
      off_analytics: Z 318, curr col 196
      def_analytics: Z 328, curr col 199
    """
    # Build current-year norms from 25-26 DMAN
    curr_norms = {}
    for row in ws_curr.iter_rows(min_row=3, values_only=True):
        name = row[player_col_curr]
        if not name or not isinstance(name, str): continue
        curr_norms[name.strip()] = {
            'overall':       row[187] if len(row)>187 else None,
            'off_prod':      row[190] if len(row)>190 else None,
            'def_prod':      row[193] if len(row)>193 else None,
            'off_analytics': row[196] if len(row)>196 else None,
            'def_analytics': row[199] if len(row)>199 else None,
        }

    HIST_STARTS = {'overall':288,'off_prod':298,'def_prod':308,'off_analytics':318,'def_analytics':328}
    players = {}
    for row in ws_hist.iter_rows(min_row=4, values_only=True):
        name = row[player_col_hist]
        if not name or not isinstance(name, str): continue
        cn = curr_norms.get(name.strip(), {})
        prog = {}
        for key, start in HIST_STARTS.items():
            out = []
            for i, s in enumerate(SEASONS):
                z = row[start+i] if start+i < len(row) else None
                out.append({'season': s, 'pct': z_to_pct(z) if is_valid(z) and str(z)!='#N/A' else None})
            curr_val = cn.get(key)
            if curr_val is not None and is_valid(curr_val):
                out[-1]['pct'] = pct(curr_val)
            prog[key] = out
        players[name.strip()] = prog
    return players

def extract_prog(ws, header_row, player_col=0):
    # Legacy shim - not used directly anymore
    pass

ANALYTICS_TIPS = {
    "On-ice offense":   "How much his team outplays opponents offensively when he's on the ice at even strength.",
    "On-ice defense":   "How well his team suppresses opponents when he's on the ice at even strength.",
    "Relative offense": "How much better the team's offensive performance is when he is on the ice vs. when his teammates are on the ice without him.",
    "Relative defense": "How much better the team's defensive performance is when he is on the ice vs. when his teammates are on the ice without him.",
}

FWD_CATS = [
    ("Usage","Time on ice","TOI Norm"),("Usage","Quality of comp.","QoC Norm"),
    ("Athletic & Compete","Speed","Speed Norm"),("Athletic & Compete","Endurance","Endurance Norm"),
    ("Athletic & Compete","Physicality","physical norm dist"),
    ("Scoring","Goalscoring","goal norm dist"),("Scoring","Finishing","shot norm"),
    ("Scoring","Shot volume","volume norm"),("Scoring","Shot quality","HD norm"),
    ("Scoring","Shot speed","Shot Speed Norm"),
    ("Offense","Passing","pass norm dist"),("Offense","Playdriving","play norm dist"),
    ("Offense","Rush offense","rush norm dist"),("Offense","Forecheck offense","forecheck norm dist"),
    ("Offense","Cycle offense","cycle norm"),("Offense","Net-front presence","Grit Norm"),
    ("Offense","Faceoffs","faceoff norm dist"),
    ("Defense","Entry defense","entry d norm"),("Defense","Exits","Exits norm dist"),
    ("Defense","Shot suppression","shot sup norm"),("Defense","Takeaways","takeaways norm"),
    ("Special teams","Power play","PP Composite Norm"),("Special teams","Penalty kill","PK Composite Norm"),
    ("Special teams","Penalty differential","penalty norm dist"),
    ("Analytics","On-ice offense","team o norm dist"),("Analytics","On-ice defense","team d norm dist"),
    ("Analytics","Relative offense","rel o norm dist"),("Analytics","Relative defense","rel d norm dist"),
]

DMAN_CATS = [
    ("Usage","Time on ice","TOI Norm"),("Usage","Quality of comp.","QoC Norm"),
    ("Athletic & Compete","Speed","Speed Norm"),("Athletic & Compete","Endurance","Endurance Norm"),
    ("Athletic & Compete","Physicality","physical norm dist"),
    ("Scoring","Goalscoring","Goal norm dist"),("Scoring","Finishing","shot norm"),
    ("Scoring","Shot volume","volume norm"),("Scoring","Shot quality","Selection norm"),
    ("Scoring","Shot speed","Shot Speed Norm"),
    ("Offense","Passing","pass norm dist"),("Offense","Playdriving","play norm dist"),
    ("Offense","Transition offense","Transition norm dist"),
    ("Offense","Off. Zone Involvement","fore/cycle norm"),
    ("Defense","Entry defense","Entry D norm dist"),("Defense","D zone exits","Exits norm dist"),
    ("Defense","Shot suppression","Shot suppression norm dist"),
    ("Defense","Takeaways","takeaway norm"),("Defense","Giveaways","Giveaway norm"),
    ("Special teams","Power play","PP Composite Norm"),("Special teams","Penalty kill","PK Composite Norm"),
    ("Special teams","Penalty differential","penalty norm dist"),
    ("Analytics","On-ice offense","team o norm"),("Analytics","On-ice defense","team d norm"),
    ("Analytics","Relative offense","rel o norm"),("Analytics","Relative defense","rel d norm"),
]

def read_cap(ws):
    cap_years = {}
    for c in range(4,13):
        yr = ws.cell(row=10,column=c).value
        cap = ws.cell(row=11,column=c).value
        if yr and cap: cap_years[int(yr)] = float(cap)
    player_z = {}
    r = 31
    while True:
        name = ws.cell(row=r,column=2).value
        if name is None: break
        cz = ws.cell(row=r,column=6).value
        mz = ws.cell(row=r,column=7).value
        if isinstance(name,str) and name.strip():
            player_z[name.strip().lower()] = {
                'current_z': float(cz) if cz else None,
                'multi_yr_z': float(mz) if mz else None,
            }
        r += 1
    return {'cap_years':cap_years,'avg_pct':float(ws.cell(row=12,column=2).value),
            'stdev_pct':float(ws.cell(row=13,column=2).value),'player_z':player_z}

def fwd_playstyle(p):
    cats = p.get('categories', {})
    scoring = {c['name']: c.get('pct') or 0 for c in cats.get('Scoring', [])}
    offense  = {c['name']: c.get('pct') or 0 for c in cats.get('Offense', [])}
    rush  = offense.get('Rush offense', 0)
    fore  = offense.get('Forecheck offense', 0)
    cycle = offense.get('Cycle offense', 0)
    nf    = offense.get('Net-front presence', 0)
    best  = max(rush, fore, cycle, nf)
    if   best == rush:  zone = 'Rush'
    elif best == fore:  zone = 'Forecheck'
    elif best == cycle: zone = 'Cycle'
    else:               zone = 'Net-Front'
    # Sniper: elite finisher, legitimate goal scorer, not primarily a shot-location player
    finishing   = scoring.get('Finishing', 0)
    goalscoring = scoring.get('Goalscoring', 0)
    shot_q      = scoring.get('Shot quality', 0)
    if finishing > 90 and goalscoring > 60 and shot_q < 90:
        return 'Sniper / ' + zone
    return zone

def dman_playstyle(p):
    prog = p.get('progression', {})
    cats = p.get('categories', {})
    def last(track):
        arr = prog.get(track, [])
        return (arr[-1].get('pct') or 0) if arr else 0
    def_score  = last('def_prod') / 100
    off_score  = last('off_analytics') / 100
    # Puckmoving uses Transition offense (Transition norm dist col) for DMAN
    trans_score = next((c.get('pct') for c in cats.get('Offense', []) if c['name'] == 'Transition offense'), None)
    play_score  = next((c.get('pct') or 0 for c in cats.get('Offense', []) if c['name'] == 'Playdriving'), 0) / 100
    puck_score  = (trans_score / 100) if trans_score is not None else play_score
    phys_score  = next((c.get('pct') or 0 for c in cats.get('Athletic & Compete', []) if c['name'] == 'Physicality'), 0) / 100
    modifier  = 'Puckmoving' if puck_score > 0.7 else ('Enforcer' if phys_score > 0.8 else '')
    two_way   = off_score > 0.6 and def_score > 0.6
    if two_way:
        label = ('Two-Way ' + modifier + ' Dman').strip().replace('  ', ' ')
    else:
        primary = 'Offensive' if off_score >= def_score else 'Defensive'
        parts   = [x for x in [modifier, primary] if x]
        label   = ' '.join(parts) + ' Dman'
    return label.strip()

def calc_contract(name, cap_hit, exp_year, params):
    key = name.strip().lower()
    zd = params['player_z'].get(key)
    if not zd or not zd['multi_yr_z']: return None
    mz = zd['multi_yr_z']
    try: exp_yr = int(exp_year) if exp_year else None
    except: exp_yr = None
    years = []
    for yr, cap_m in sorted(params['cap_years'].items()):
        impl = cap_m * (params['avg_pct'] + params['stdev_pct'] * mz) * 1_000_000
        on = exp_yr is not None and yr <= exp_yr
        years.append({'year':yr,'implied_value':dollars(impl),'cap_hit':dollars(cap_hit) if on else None,
                      'surplus':dollars(impl-cap_hit) if on else None,'on_contract':on})
    cy = [y for y in years if y['on_contract']]
    surps = [y['surplus'] for y in cy if y['surplus'] is not None]
    return {'current_z':round(zd['current_z'],4) if zd['current_z'] else None,
            'multi_yr_z':round(mz,4),'years':years,
            'avg_aav_excess':dollars(sum(surps)/len(surps)) if surps else None,
            'sum_surplus':dollars(sum(surps)) if surps else None,
            'sum_value':dollars(sum(y['implied_value'] for y in cy)) if cy else None}

def extract_players(ws, cats, ovl_col, ps, toi_map, prog_map, is_dman=False):
    header = list(ws.iter_rows(min_row=2, max_row=2, values_only=True))[0]
    col = build_header(header)
    cat_idx = [(g, d, col.get(s)) for g,d,s in cats]
    players = []
    for row in ws.iter_rows(min_row=3, values_only=True):
        name = row[col.get('Player',1)] if 'Player' in col else None
        if not name or not isinstance(name, str): continue
        def get(c): idx=col.get(c); return row[idx] if idx is not None and idx<len(row) else None
        cats_out = {}
        for g,d,idx in cat_idx:
            if g not in cats_out: cats_out[g] = []
            val = row[idx] if idx is not None and idx<len(row) else None
            cats_out[g].append({'name':d,'pct':pct(val),'tip':ANALYTICS_TIPS.get(d)})
        gp = get('GP')
        toi_str = fmt_toi(toi_map.get(name.strip()), gp)
        pos = safe_str(get('POS'))
        if is_dman and pos == 'D':
            handed = safe_str(get('HANDED'))
            if handed and 'Left' in handed: pos = 'LD'
            elif handed and 'Right' in handed: pos = 'RD'
        ovl_raw = row[col[ovl_col]] if ovl_col in col and col[ovl_col]<len(row) else None
        players.append({
            'name':safe_str(name),'team':safe_str(get('TEAM')),'pos':pos,
            'age':round(get('Age'),1) if isinstance(get('Age'),float) else get('Age'),
            'gp':gp,'toi_per_gp':toi_str,'height':fmt_height(get('Height (ft)')),
            'weight':int(get('Weight (lbs)')) if get('Weight (lbs)') else None,
            'overall_pct':pct(ovl_raw),'cap':dollars(get('Cap')),
            'value':dollars(get('Value')),'surplus':dollars(get('Surplus Value')),
            'role':safe_str(get('Role')),'actual_role':safe_str(get('Actual Role')),
            'exp_year':safe_str(get('EXP. YEAR')),'expiry_type':safe_str(get('EXPIRY')),
            'playstyle':None,'avg_annual_surplus':None,'contract_calc':None,
            'progression':prog_map.get(name.strip()),'categories':cats_out,
            'pim':pim_map.get(name.strip()),
            'prior_season': name.strip() in prior_season_players,
        })
    return players

# Open workbooks
wb  = openpyxl.load_workbook(VALUE_DRIVE, read_only=True, data_only=True, keep_vba=False)
db  = openpyxl.load_workbook(DATABASE,    read_only=True, data_only=True, keep_vba=False)

# TOI from EDGE
toi_map = {}
for row in db['EDGE'].iter_rows(min_row=3, values_only=True):
    n, t = row[0], row[1]
    if isinstance(n,str) and n.strip() and isinstance(t,(int,float)) and t>0:
        toi_map[n.strip()] = t
print(f"  TOI data: {len(toi_map)} players")

# Raw PIM from Database (for Lady Byng)
# FWD: indiv fwd col 54 = PIM; DMAN: indiv dmen col 51 = Total Penalties
pim_map = {}
for row in db['indiv fwd'].iter_rows(min_row=5, values_only=True):
    n, v = row[2], row[54] if len(row) > 54 else None
    if isinstance(n, str) and isinstance(v, (int, float)):
        pim_map[n.strip()] = int(v)
for row in db['indiv dmen'].iter_rows(min_row=5, values_only=True):
    n, v = row[2], row[51] if len(row) > 51 else None
    if isinstance(n, str) and isinstance(v, (int, float)):
        pim_map[n.strip()] = int(v)
print(f"  PIM data: {len(pim_map)} players")

# Prior-season player sets for Calder eligibility
# Ineligible if they appear in 24-25 FWD or 24-25 DMAN
prior_season_players = set()
for sheet_name, name_col in [('2024-25 FWD', 1), ('24-25 DMAN', 1)]:
    try:
        ws_prior = wb[sheet_name]
        for row in ws_prior.iter_rows(min_row=3, values_only=True):
            n = row[name_col] if len(row) > name_col else None
            if isinstance(n, str) and n.strip():
                prior_season_players.add(n.strip())
    except KeyError:
        pass
print(f"  Prior season players: {len(prior_season_players)}")

fwd_ps   = None  # playstyle now computed from category data post-extraction
dman_ps  = None
fwd_cap  = read_cap(wb['FWD CAP'])
dman_cap = read_cap(wb['D CAP'])

fwd_prog  = extract_prog_fwd(wb['FWD 25'], wb['25-26 FWD'])
dman_prog = extract_prog_dman(wb['DMAN 24'], wb['25-26 DMAN'])
print(f"  Progression: {len(fwd_prog)} FWD, {len(dman_prog)} DMAN")

fwd_players = extract_players(wb['25-26 FWD'], FWD_CATS, 'Overall Norm', fwd_ps, toi_map, fwd_prog)
for p in fwd_players:
    cc = calc_contract(p['name'],p['cap'],p['exp_year'],fwd_cap)
    p['contract_calc'] = cc
    if cc: p['avg_annual_surplus'] = cc['avg_aav_excess']
    p['playstyle'] = fwd_playstyle(p)
print(f"  Forwards: {len(fwd_players)}")

dman_players = extract_players(wb['25-26 DMAN'], DMAN_CATS, 'Overall Norm', dman_ps, toi_map, dman_prog, is_dman=True)
for p in dman_players:
    cc = calc_contract(p['name'],p['cap'],p['exp_year'],dman_cap)
    p['contract_calc'] = cc
    if cc: p['avg_annual_surplus'] = cc['avg_aav_excess']
    p['playstyle'] = dman_playstyle(p)
print(f"  Defensemen: {len(dman_players)}")

# ── Goalies from G 24 ────────────────────────────────────────────────────────
print("  Extracting goalies...")
def fmt_height_g(v):
    if not isinstance(v, (int,float)): return None
    f = int(v); i = round((v-f)*12)
    return str(f) + "'" + str(i) + '"'

goalies = []
for row in wb['G 24'].iter_rows(min_row=4, values_only=True):
    name = row[0]
    if not isinstance(name, str) or not name.strip() or name == 'Player': continue
    team = str(row[1]).strip() if row[1] else None
    cap = row[14] if isinstance(row[14], (int,float)) else None
    value = row[15] if isinstance(row[15], (int,float)) else None
    surplus = row[16] if isinstance(row[16], (int,float)) else None
    exp_year = str(int(row[17])) if isinstance(row[17], (int,float)) else (str(row[17]) if row[17] else None)
    expiry = str(row[18]).strip() if row[18] else None
    value_2526 = row[19] if isinstance(row[19], (int,float)) else None
    gp = int(row[5]) if isinstance(row[5], (int,float)) else None
    age = round(row[4], 1) if isinstance(row[4], (int,float)) else None
    current_t = team.split(',')[-1].strip() if team and ',' in team else team
    comp_norm = row[13] if isinstance(row[13], (int,float)) else None
    goalies.append({
        'name': name.strip(), 'team': team, 'current_team': current_t,
        'pos': 'G', 'age': age, 'gp': gp,
        'height': fmt_height_g(row[2]),
        'weight': int(row[3]) if isinstance(row[3], (int,float)) else None,
        'cap': int(cap) if cap else None,
        'value': round(value) if value else None,
        'value_2526': round(value_2526) if value_2526 else None,
        'surplus': round(surplus) if surplus else None,
        'comp_value_norm': round(comp_norm * 100) if comp_norm is not None else None,
        'exp_year': exp_year, 'expiry_type': expiry, '_type': 'G',
    })
print(f"  Goalies: {len(goalies)}")

wb.close(); db.close()

with open(OUT / 'players_fwd.json',  'w') as f: json.dump(fwd_players,  f, indent=2, default=str)
with open(OUT / 'players_dman.json', 'w') as f: json.dump(dman_players, f, indent=2, default=str)
print("  Saved players_fwd.json + players_dman.json")

# ════════════════════════════════════════════════════════════════════════════════
# STEP 2: Extract comp vectors
# ════════════════════════════════════════════════════════════════════════════════
print()
print("Step 2/3 — Extracting comp vectors from Comp Sheet...")

def safe_float(v):
    if v is None or not isinstance(v,(int,float)) or math.isnan(v): return 0.0
    return round(float(v),5)

def make_vec(row, col_map, names):
    return [safe_float(row[col_map[c]]) if c in col_map and col_map[c]<len(row) else 0.0 for c in names]

FWD_INDIVIDUAL = [
    'Height (ft)','Age','TOI/GP','Goals','First Assists','Second Assists','IPP','SH%','ixG','iFF','iSCF','iHDCF',
    'Rush Attempts','Rebounds Created','PIM','Penalties Drawn','Giveaways','Takeaways','Hits','Hits Taken',
    'Shots Blocked','SHOOTING ON UNBLOCKED SHOTS ABOVE EXPECTED','NET MISS % Above Expected','SHOOTING TALENT ABOVE AVG',
    'Shot Assists/60','Chance Assists/60','Chance Contributions/60','HD Passes/60','One-timer Assists/60',
    'Rush Offense/60','Forecheck Offense/60','Pressures/60','Recoveries/60','Cycle Offense/60','One-timer/60',
    'Gritty Shots','Shot off HD/60','Entries/60','Carry%','Carry Chance%','Entries w/ Chances per 60',
    'Failed Entries %','Carries Against per 60','Carry Against%','Denials per 60','Denials %',
    'Chances Allowed per 60','Chance Against%','DZ Puck Touches per 60','Retrievals per 60',
    'Successful Retrieval%','Retrievals Leading to Exits per 60','Missed Pass/DZ Exchange','Exit off Retrieval%',
    'Exits per 60','Exits w/ Possession per 60','Failed Exit%','DZ Assists/60','Assisted Exits/60',
    'Carried Exits/60','Clears/60','DZ Counters/60','DZ Controlled Breakout/60','Secondary Assists/60',
    '5v5 D WAR','5v5 O WAR','Speed Z','Endurance Z','Shot Speed Z','Quality of Competition',
    'Weight (lbs)','shot z','REL O Z','REL D Z','PP Composite','PK Composite',
]
FWD_MICROSTAT = [
    'Height (ft)','Age','TOI/GP','Rush Attempts','Rebounds Created','PIM','Penalties Drawn',
    'Giveaways','Takeaways','Hits','Hits Taken','Shots Blocked',
    'SHOOTING ON UNBLOCKED SHOTS ABOVE EXPECTED','NET MISS % Above Expected','SHOOTING TALENT ABOVE AVG',
    'Shot Assists/60','Chance Assists/60','Chance Contributions/60','HD Passes/60','One-timer Assists/60',
    'Rush Offense/60','Forecheck Offense/60','Pressures/60','Recoveries/60','Cycle Offense/60','One-timer/60',
    'Gritty Shots','Shot off HD/60','Entries/60','Carry%','Carry Chance%','Entries w/ Chances per 60',
    'Failed Entries %','Exits per 60','Exits w/ Possession per 60','Failed Exit%',
    'DZ Assists/60','DZ Controlled Breakout/60','Secondary Assists/60',
    'Speed Z','Endurance Z','Shot Speed Z','Weight (lbs)',
]
FWD_STYLE = [
    'Speed Z','Endurance Z','Shot Speed Z','Finishing Z','Pass Z','Play Z','Volume Shooting',
    'selection z','Rush Z','Forecheck Z','cycle z','Grit','Phys Z','Faceoff Z','Exits Z','shot z',
]
DMAN_INDIVIDUAL = [
    'Height (ft)','Age','TOI/GP','Goals','First Assists','Second Assists','IPP','SH%','ixG','iFF','iSCF','iHDCF',
    'Rush Attempts','Rebounds Created','PIM','Penalties Drawn','Giveaways','Takeaways','Hits','Hits Taken',
    'Shots Blocked','SHOOTING ON UNBLOCKED SHOTS ABOVE EXPECTED','NET MISS % Above Expected','SHOOTING TALENT ABOVE AVG',
    'Shot Assists/60','Chance Assists/60','Chance Contributions/60','HD Passes/60','One-timer Assists/60',
    'Rush Offense/60','Forecheck Offense/60','Pressures/60','Recoveries/60','Cycle Offense/60','One-timer/60',
    'Gritty Shots','Shot off HD/60','Entries/60','Carry%','Carry Chance%','Entries w/ Chances per 60',
    'Failed Entries %','Carries Against per 60','Carry Against%','Denials per 60','Denial%',
    'Chances Allowed per 60','Chance Against%','DZ Puck Touches per 60','Retrievals per 60',
    'Successful Retrieval%','Retrievals Leading to Exits per 60','Missed Pass/DZ Exchange','Exit off Retrieval%',
    'Exits per 60','Exits w/ Possession per 60','Failed Exit%','DZ Assists/60','Assisted Exits/60',
    'Carried Exits/60','Clears/60','DZ Counters/60','DZ Controlled Breakout/60','Secondary Assists/60',
    '5v5 D WAR','5v5 O WAR','Speed Z','Endurance Z','Shot Speed Z',
    'QoC','Weight (lbs)','suppress Z','REL O Z','REL D Z','PP Composite','PK Composite',
]
DMAN_MICROSTAT = [
    'Rush Attempts','Rebounds Created','PIM','Penalties Drawn','Giveaways','Takeaways','Hits','Hits Taken',
    'Shots Blocked','SHOOTING ON UNBLOCKED SHOTS ABOVE EXPECTED','NET MISS % Above Expected','SHOOTING TALENT ABOVE AVG',
    'Shot Assists/60','Chance Assists/60','Chance Contributions/60','HD Passes/60','One-timer Assists/60',
    'Rush Offense/60','Forecheck Offense/60','Pressures/60','Recoveries/60','Cycle Offense/60','One-timer/60',
    'Gritty Shots','Shot off HD/60','Entries/60','Carry%','Carry Chance%','Entries w/ Chances per 60',
    'Failed Entries %','Carries Against per 60','Carry Against%','Denials per 60','Denial%',
    'Chances Allowed per 60','Chance Against%','DZ Puck Touches per 60','Retrievals per 60',
    'Successful Retrieval%','Retrievals Leading to Exits per 60','Missed Pass/DZ Exchange','Exit off Retrieval%',
    'Exits per 60','Exits w/ Possession per 60','Failed Exit%','DZ Assists/60','Assisted Exits/60',
    'Carried Exits/60','Clears/60','DZ Counters/60','DZ Controlled Breakout/60','Secondary Assists/60',
    'Speed Z','Endurance Z','Shot Speed Z',
]
DMAN_STYLE = [
    'Speed Z','Endurance Z','Shot Speed Z','Finishing Z','Pass Z','Play Z',
    'Volume Shooting','Transition Z','fore/cycle z','Phys Z',
    'Entry D Z','Exits Z','suppress Z','PP Composite','PK Composite',
]

def extract_comp_sheet(ws, name_col, disp_col, team_col, pos_col, hdr_idx,
                       indiv, micro, style, fix_base=False):
    col = {}
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == hdr_idx:
            for j, v in enumerate(row):
                if v is not None: col[str(v).strip()] = j
            break
    records = []
    for row in ws.iter_rows(min_row=hdr_idx+2, values_only=True):
        base = row[name_col] if name_col < len(row) else None
        disp = row[disp_col] if disp_col < len(row) else None
        if not isinstance(base,str) or not isinstance(disp,str): continue
        parts = disp.strip().split()
        yr = int(parts[-1]) if parts and parts[-1].isdigit() and len(parts[-1])==4 else None
        if yr is None: continue
        base_clean = re.sub(r'\s+\d{4}$','',base.strip()).strip() if fix_base else base.strip()
        team = row[team_col] if team_col < len(row) else None
        pos  = row[pos_col]  if pos_col  < len(row) else None
        records.append({
            'display_name': disp.strip(),
            'base_name': base_clean,
            'year': yr,
            'team': str(team).strip() if team else None,
            'pos':  str(pos).strip()  if pos  else None,
            'individual_vec': make_vec(row, col, indiv),
            'microstat_vec':  make_vec(row, col, micro),
            'style_vec':      make_vec(row, col, style),
        })
    return records

wb2 = openpyxl.load_workbook(COMP_SHEET, read_only=True, data_only=True)
fwd_comps  = extract_comp_sheet(wb2['FWD'],  0,1,2,3, 1, FWD_INDIVIDUAL,  FWD_MICROSTAT,  FWD_STYLE,  False)
dman_comps = extract_comp_sheet(wb2['DMAN'], 2,1,3,4, 1, DMAN_INDIVIDUAL, DMAN_MICROSTAT, DMAN_STYLE, True)
wb2.close()

print(f"  FWD comps: {len(fwd_comps)} player-seasons")
print(f"  DMAN comps: {len(dman_comps)} player-seasons")

with open(OUT / 'comp_fwd.json',  'w') as f: json.dump(fwd_comps,  f, separators=(',',':'))
with open(OUT / 'comp_dman.json', 'w') as f: json.dump(dman_comps, f, separators=(',',':'))
print("  Saved comp_fwd.json + comp_dman.json")

# ════════════════════════════════════════════════════════════════════════════════
# STEP 3: Build embedded HTML
# ════════════════════════════════════════════════════════════════════════════════
print()
print("Step 3/3 — Building embedded HTML...")

with open(HTML_TPL) as f: html = f.read()

# Minify all data
fwd_min   = json.dumps(fwd_players,  separators=(',',':'))
dman_min  = json.dumps(dman_players, separators=(',',':'))
cfwd_min  = json.dumps(fwd_comps,    separators=(',',':'))
cdman_min = json.dumps(dman_comps,   separators=(',',':'))

# Add current_team (last in comma-separated list = post-trade current team)
for p in fwd_players + dman_players:
    team = p.get('team') or ''
    p['current_team'] = team.split(',')[-1].strip() if team else None

# Embed player data
html = html.replace('const _FWD_DATA = PLACEHOLDER_FWD;',  f'const _FWD_DATA = {fwd_min};')
html = html.replace('const _DMAN_DATA = PLACEHOLDER_DMAN;', f'const _DMAN_DATA = {dman_min};')

# Embed comp data into loadData block
# Write JSON data files separately (fetched async - much faster load)
with open(OUT / 'players_fwd.json',  'w') as f: f.write(fwd_min)
with open(OUT / 'players_dman.json', 'w') as f: f.write(dman_min)
goalies_min = json.dumps(goalies, separators=(',',':'))
with open(OUT / 'players_goalie.json', 'w') as f: f.write(goalies_min)
with open(OUT / 'comp_fwd.json',     'w') as f: f.write(cfwd_min)
with open(OUT / 'comp_dman.json',    'w') as f: f.write(cdman_min)

# Template uses fetch() placeholders - no data embedded in HTML
html = html.replace('PLACEHOLDER_FWD',       '"players_fwd.json"')
html = html.replace('PLACEHOLDER_DMAN',      '"players_dman.json"')
html = html.replace('PLACEHOLDER_COMP_FWD',  '"comp_fwd.json"')
html = html.replace('PLACEHOLDER_COMP_DMAN', '"comp_dman.json"')
html = html.replace('PLACEHOLDER_GOALIES',   '"players_goalie.json"')

with open(OUT / 'nhl-analytics-v3-embedded.html', 'w') as f: f.write(html)

# ── Summary ───────────────────────────────────────────────────────────────────
total_kb = sum((OUT/fn).stat().st_size for fn in ['players_fwd.json','players_dman.json','comp_fwd.json','comp_dman.json','nhl-analytics-v3-embedded.html']) // 1024
print(f"  HTML: {(OUT/'nhl-analytics-v3-embedded.html').stat().st_size//1024}KB")
print(f"  Comp files: {(OUT/'comp_fwd.json').stat().st_size//1024}KB + {(OUT/'comp_dman.json').stat().st_size//1024}KB")
print(f"  Total output: {total_kb}KB")
print()
print("Done! Files are in ./output/")
print()
print("To use locally:  Open nhl-analytics-v3-embedded.html + keep comp_fwd.json and comp_dman.json in the same folder")
print("To deploy:       Drag the output/ folder onto your Netlify site")
